from openpyxl import load_workbook
import json

wb = load_workbook(filename= 'dev.xlsx')

sheets = wb.sheetnames
sheetsData = []
dataDuplicateCount = {}

for sheet in sheets:
  sheetsData.append(wb[sheet])

for sheet in sheetsData:
  dataDuplicateCount[sheet.title] = {}
  for index in range(2, sheet.max_row) :
    if str(sheet['A' + str(index)].value) in dataDuplicateCount[sheet.title]:
      dataDuplicateCount[sheet.title][str(sheet['A' + str(index)].value)]+=1
    else :
      dataDuplicateCount[sheet.title][str(sheet['A' + str(index)].value)] = 1

print(json.dumps(dataDuplicateCount, indent=4, sort_keys=True))

# while sheet is not None :
#   sheets.append(sheet)
#   sheet = wb['Sheet' + str(++index)]

# for sh in sheets:
#   print(sh['A2'].value)
